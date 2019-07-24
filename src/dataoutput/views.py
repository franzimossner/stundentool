from django.shortcuts import render, get_object_or_404
from datainput.models import Schulklasse, Raum, Tag, Stunde, Lehrer, Schulfach, Uebergreifung, VorgabeEinheit, Lehreinheit, OptimierungsErgebnis
from django.contrib.auth.decorators import login_required

# Fuer Excel export
import xlwt
import openpyxl
from django.http import HttpResponse
# Fuer PDF export
import reportlab
import io
from django.http import FileResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle,Paragraph
from reportlab.lib.enums import TA_JUSTIFY, TA_LEFT, TA_CENTER


''' Hier sind die Hauptviews zum Darstellen der dataoutput Seiten
'''
@login_required
def output_main(request):
    return render(request, 'dataoutput/output_main.html',)

@login_required
def output_teacher(request):
    run = OptimierungsErgebnis.objects.order_by('-Index').first()
    lehrers = Lehrer.objects.order_by('Name')
    return render(request, 'dataoutput/output_teacher.html', {'lehrers': lehrers, 'run': run})

@login_required
def output_classes(request):
    # aktuell nur die neueste Version zum download für alle Klassen, muss noch in eine Tabelle umgewandelt werden
    run = OptimierungsErgebnis.objects.order_by('-Index').first()
    klassen = Schulklasse.objects.order_by('Name')
    return render(request, 'dataoutput/output_classes.html', {'klassen': klassen, 'run': run})

@login_required
def class_detail(request, klasse):
    klasse = get_object_or_404(Schulklasse, Name=klasse)
    runs = OptimierungsErgebnis.objects.all()
    tage = Tag.objects.order_by('Index')
    stunden = Stunde.objects.order_by('Index')
    runToStundenplan = []
    for run in runs:
        stundenplan = []
        for stunde in stunden:
            tagesplan = []
            for tag in tage:
                tagesplan.append(klasse.lehreinheit_set.filter(Zeitslot__Stunde=stunde, Zeitslot__Tag=tag, run=run).all())
                #tagesplan.append(klasse.lehreinheit_set.filter(Zeitslot__Stunde=stunde, Zeitslot__Tag=tag, run=run).first())
            stundenplan.append((stunde, tagesplan))
        runToStundenplan.append((klasse, stundenplan, run))
    return render(request, 'dataoutput/class_detail.html', {'runToStundenplan': runToStundenplan ,
    'tage': tage ,
    'stunden': stunden,
    'klasse': klasse,
    'runs': runs
    })

@login_required
def teacher_detail(request, lehrer):
    lehrer = get_object_or_404(Lehrer, Kurzname=lehrer)
    runs = OptimierungsErgebnis.objects.all()
    tage = Tag.objects.order_by('Index')
    stunden = Stunde.objects.order_by('Index')
    runToStundenplan = []
    for run in runs:
        stundenplan = []
        for stunde in stunden:
            tagesplan = []
            for tag in tage:
                tagesplan.append(lehrer.lehreinheit_set.filter(Zeitslot__Stunde=stunde, Zeitslot__Tag=tag, run=run).first())
            stundenplan.append((stunde, tagesplan))
        runToStundenplan.append((lehrer, stundenplan, run))
    return render(request, 'dataoutput/teacher_detail.html', {'runToStundenplan': runToStundenplan,
    'lehrer':lehrer,
    'tage': tage,
    'stunden': stunden,
    'runs' : runs
    })



''' Ab hier Code zum Download von Klassenstundenplänen in Excel
'''

@login_required
def download_excel_data_Alleklassen(request, run):
    # content-type of response
    response = HttpResponse(content_type='application/ms-excel')
    #decide file name
    response['Content-Disposition'] = 'attachment; filename="alleKlassen_{0}_.xls"'.format(run)
    #creating workbook
    wb = xlwt.Workbook(encoding='utf-8')

    #adding sheet for each class
    klassen = Schulklasse.objects.order_by('Name')
    tage = Tag.objects.order_by('Index')
    stunden = Stunde.objects.order_by('Index')
    Run = OptimierungsErgebnis.objects.get(Index=run)

    for klasse in klassen:
        ws = wb.add_sheet("Klasse {0} ".format(klasse))
        row_num = 0

        font_style = xlwt.XFStyle()
        style = xlwt.XFStyle()
        style.alignment.wrap = 1
        # headers are bold
        font_style.font.bold = True
        #column header names
        columns = ['Stunden']
        for tag in tage:
               columns.append(tag.Tag)
        #write column headers in sheet
        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num], font_style)

        for stunde_num, stunde in enumerate(stunden):
            ws.write(stunde_num + 1, 0, stunde.Stunde, font_style)
            colwidth = max(len(str(stunde.Stunde)) for stunde in stunden)
            ws.col(0).width = colwidth * 367

        # Sheet body, remaining rows
        font_style = xlwt.XFStyle()

        for stunde in stunden:
            for tag in tage:
                celldata= klasse.lehreinheit_set.filter(Zeitslot__Stunde=stunde, Zeitslot__Tag=tag, run=Run).all()
                cellobject = ""
                for lehreinheit in celldata:
                    cellobject += "{0} ({1})".format(lehreinheit.Schulfach.Name, lehreinheit.Lehrer.Kurzname)
                    cellobject += "\n"

                colwidth = max(len(cellobject), len(str(tag))) + 3
                row_num = stunde.Index
                tag_num = tag.Index

                # Warum 367??
                # Default value of width is 2962 units and excel points it to as 8.11 units. Hence i am multiplying 367 to length of data.
                ws.col(tag_num).width = colwidth * 367
                # ws.write(row_num, col_num, content, font_style)
                ws.write(row_num, tag_num , cellobject, style)
    wb.save(response)
    return response

@login_required
def download_excel_1klasse_1run(request, klasse, run):
    # content-type of response
    response = HttpResponse(content_type='application/ms-excel')
    #decide file name
    response['Content-Disposition'] = 'attachment; filename="klassenplan_{0}_{1}.xls"'.format(klasse, run)
    #creating workbook
    wb = xlwt.Workbook(encoding='utf-8')

    #adding sheet for the class
    tage = Tag.objects.order_by('Index')
    stunden = Stunde.objects.order_by('Index')
    run = OptimierungsErgebnis.objects.get(Index=run)

    ws = wb.add_sheet("Klasse {0} ".format(klasse))
    row_num = 0

    font_style = xlwt.XFStyle()
    style = xlwt.XFStyle()
    style.alignment.wrap = 1
    # headers are bold
    font_style.font.bold = True
    #column header names
    columns = ['Stunden']
    for tag in tage:
           columns.append(tag.Tag)
    #write column headers in sheet
    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style)

    for stunde_num, stunde in enumerate(stunden):
        ws.write(stunde_num + 1, 0, stunde.Stunde, font_style)
        colwidth = max(len(str(stunde.Stunde)) for stunde in stunden)
        ws.col(0).width = colwidth * 367

    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()
    #get data to filter the database
    Klasse= Schulklasse.objects.get(Name=klasse)

    for stunde in stunden:
        for tag in tage:
            celldata= Klasse.lehreinheit_set.filter(Zeitslot__Stunde=stunde, Zeitslot__Tag=tag, run=run).all()
            cellobject = ""
            for lehreinheit in celldata:
                cellobject += "{0} ({1})".format(lehreinheit.Schulfach.Name, lehreinheit.Lehrer.Kurzname)
                cellobject += "\n"

            colwidth = max(len(cellobject), len(str(tag))) + 3
            row_num = stunde.Index
            tag_num = tag.Index

            # Warum 367??
            # Default value of width is 2962 units and excel points it to as 8.11 units. Hence i am multiplying 367 to length of data.
            ws.col(tag_num).width = colwidth * 367
            # ws.write(row_num, col_num, content, font_style)
            ws.write(row_num, tag_num , cellobject, style)

    wb.save(response)
    return response



''' Ab hier Code zum Download von Lehrerstundenplänen in Excel
'''

@login_required
def download_excel_1lehrer_1run(request, lehrer, run):
    # content-type of response
    response = HttpResponse(content_type='application/ms-excel')
    #decide file name
    response['Content-Disposition'] = 'attachment; filename="lehrerplan_{0}_{1}.xls"'.format(lehrer, run)
    #creating workbook
    wb = xlwt.Workbook(encoding='utf-8')

    #adding sheet for the class
    tage = Tag.objects.order_by('Index')
    stunden = Stunde.objects.order_by('Index')
    run = OptimierungsErgebnis.objects.get(Index=run)

    ws = wb.add_sheet("{0} ".format(lehrer))
    # Row to choose for the header of the data
    row_num = 0

    font_style = xlwt.XFStyle()
    style = xlwt.XFStyle()
    style.alignment.wrap = 1
    # headers are bold
    font_style.font.bold = True
    #column header names
    columns = ['Stunden']
    for tag in tage:
           columns.append(tag.Tag)
    #write column headers in sheet
    for col_num in range(len(columns)):
        ws.write(row_num, col_num, columns[col_num], font_style)

    for stunde_num, stunde in enumerate(stunden):
        ws.write(stunde_num + 1, 0, stunde.Stunde, font_style)
        colwidth = max(len(str(stunde.Stunde)) for stunde in stunden)
        ws.col(0).width = colwidth * 367

    # Sheet body, remaining rows
    font_style = xlwt.XFStyle()
    #get data to filter the database
    Lehrperson = Lehrer.objects.get(Kurzname=lehrer)

    for stunde in stunden:
        for tag in tage:
            celldata= Lehrperson.lehreinheit_set.filter(Zeitslot__Stunde=stunde, Zeitslot__Tag=tag, run=run).all()
            cellobject = ""
            for lehreinheit in celldata:
                cellobject += "{0} ({1})".format(lehreinheit.Schulfach.Name, lehreinheit.Klasse.Name)
                cellobject += "\n"

            colwidth = max(len(cellobject), len(str(tag))) + 3
            row_num = stunde.Index
            tag_num = tag.Index

            # Warum 367??
            # Default value of width is 2962 units and excel points it to as 8.11 units. Hence i am multiplying 367 to length of data.
            ws.col(tag_num).width = colwidth * 367
            # ws.write(row_num, col_num, content, font_style)
            ws.write(row_num, tag_num , cellobject, style)

    wb.save(response)
    return response


@login_required
def download_excel_data_Allelehrer(request, run):
    # content-type of response
    response = HttpResponse(content_type='application/ms-excel')
    #decide file name
    response['Content-Disposition'] = 'attachment; filename="alleLehrer_{0}_.xls"'.format(run)
    #creating workbook
    wb = xlwt.Workbook(encoding='utf-8')

    #adding sheet for each class
    lehrers = Lehrer.objects.order_by('Name')
    tage = Tag.objects.order_by('Index')
    stunden = Stunde.objects.order_by('Index')
    run = OptimierungsErgebnis.objects.get(Index=run)

    for lehrer in lehrers:
        ws = wb.add_sheet("{0} ".format(lehrer))
        # Row to choose for the header of the data
        row_num = 0

        font_style = xlwt.XFStyle()
        style = xlwt.XFStyle()
        style.alignment.wrap = 1
        # headers are bold
        font_style.font.bold = True
        #column header names
        columns = ['Stunden']
        for tag in tage:
               columns.append(tag.Tag)
        #write column headers in sheet
        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num], font_style)

        for stunde_num, stunde in enumerate(stunden):
            ws.write(stunde_num + 1, 0, stunde.Stunde, font_style)
            colwidth = max(len(str(stunde.Stunde)) for stunde in stunden)
            ws.col(0).width = colwidth * 367

        # Sheet body, remaining rows
        font_style = xlwt.XFStyle()

        for stunde in stunden:
            for tag in tage:
                celldata= lehrer.lehreinheit_set.filter(Zeitslot__Stunde=stunde, Zeitslot__Tag=tag, run=run).all()
                cellobject = ""
                for lehreinheit in celldata:
                    cellobject += "{0} ({1})".format(lehreinheit.Schulfach.Name, lehreinheit.Klasse.Name)
                    cellobject += "\n"

                colwidth = max(len(cellobject), len(str(tag))) + 3
                row_num = stunde.Index
                tag_num = tag.Index

                # Warum 367??
                # Default value of width is 2962 units and excel points it to as 8.11 units. Hence i am multiplying 367 to length of data.
                ws.col(tag_num).width = colwidth * 367
                # ws.write(row_num, col_num, content, font_style)
                ws.write(row_num, tag_num , cellobject, style)
    wb.save(response)
    return response


''' Ab hier Code zum Download von Klassenstundenplänen als PDF
'''

#
def download_pdf_data_Alleklassen(request, run):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment;   filename="alleKlassen_{0}.pdf" '.format(run)

    # buffer
    buffer = io.BytesIO()
    # create pdf object
    p = canvas.Canvas(buffer, pagesize=landscape(A4))
    #width, height = A4
    styles =getSampleStyleSheet()
    styleBH = styles["Normal"]
    styleBH.alignment = TA_CENTER

    # Insert the content to the pdf, done with Reportlab documentation
    klassen = Schulklasse.objects.order_by('Name')
    lehrers = Lehrer.objects.order_by('Name')
    tage = Tag.objects.order_by('Index')
    stunden = Stunde.objects.order_by('Index')
    Run = OptimierungsErgebnis.objects.get(Index=run)

    for klasse in klassen:
        data= [['Stunden']]
        for tag in tage:
            data[0].append(str(tag.Tag).encode('utf-8'))
        for stunde in stunden:
            data.append([str(stunde.Stunde).encode('utf-8')])
            for tag in tage:
                data[-1].append("")

        #dataunits = Lehreinheit.objects.filter(Klasse=klasse, run=Run)
        for stunde in stunden:
            for tag in tage:
                celldata= klasse.lehreinheit_set.filter(Zeitslot__Stunde=stunde, Zeitslot__Tag=tag, run=Run).all()
                cellobject = ""
                for lehreinheit in celldata:
                    row_num = stunde.Index
                    tag_num = tag.Index
                    cellobject += "{0} ({1})\n".format(lehreinheit.Schulfach.Name, lehreinheit.Lehrer.Kurzname)
                    data[row_num][tag_num] = cellobject

        width, height = A4
        table = Table(data, hAlign='RIGHT')
        table.setStyle(TableStyle([
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('BACKGROUND',(0,0),(-1,0),colors.lavender),
                        ('BACKGROUND',(0,0),(0,-1),colors.lavender),
                        ('INNERGRID',(0,0),(-1,-1),0.25, colors.black)
                        ]))
        # table.wrapOn(p, width, height)
        table.wrapOn(p, width, height)
        table.drawOn(p, 5*cm, 6*cm)
        table.hAlign='TA_CENTER'
        #p.Table(data)
        #t.setStyle(tblStyle)
        p.showPage()

    p.save()
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response


def download_pdf_1klasse_1run(request, klasse, run):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment;   filename="Klasse{0}_{1}.pdf" '.format(klasse,run)

    # buffer
    buffer = io.BytesIO()
    # create pdf object
    p = canvas.Canvas(buffer, pagesize=landscape(A4))
    #width, height = A4
    styles =getSampleStyleSheet()
    styleBH = styles["Normal"]
    styleBH.alignment = TA_CENTER

    # Insert the content to the pdf, done with Reportlab documentation
    klassen = Schulklasse.objects.order_by('Name')
    lehrers = Lehrer.objects.order_by('Name')
    tage = Tag.objects.order_by('Index')
    stunden = Stunde.objects.order_by('Index')
    Run = OptimierungsErgebnis.objects.get(Index=run)

    data= [['Stunden']]
    for tag in tage:
        data[0].append(str(tag.Tag).encode('utf-8'))
    for stunde in stunden:
        data.append([str(stunde.Stunde).encode('utf-8')])
        for tag in tage:
            data[-1].append("")

    Klasse = Schulklasse.objects.get(Name=klasse)
    for stunde in stunden:
        for tag in tage:
            celldata= Klasse.lehreinheit_set.filter(Zeitslot__Stunde=stunde, Zeitslot__Tag=tag, run=Run).all()
            cellobject = ""
            for lehreinheit in celldata:
                row_num = stunde.Index
                tag_num = tag.Index
                cellobject += "{0} ({1})\n".format(lehreinheit.Schulfach.Name, lehreinheit.Lehrer.Kurzname)
                data[row_num][tag_num] = cellobject

    width, height = A4
    table = Table(data, hAlign='RIGHT')
    table.setStyle(TableStyle([
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('BACKGROUND',(0,0),(-1,0),colors.lavender),
                    ('BACKGROUND',(0,0),(0,-1),colors.lavender),
                    ('INNERGRID',(0,0),(-1,-1),0.25, colors.black)
                    ]))
    # table.wrapOn(p, width, height)
    table.wrapOn(p, width, height)
    table.drawOn(p, 5*cm, 6*cm)
    table.hAlign='TA_CENTER'
    #p.Table(data)
    #t.setStyle(tblStyle)
    p.showPage()

    p.save()
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response
